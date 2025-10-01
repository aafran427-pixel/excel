from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Profit-Loss Sheet"

# Define the headers and structure of the sheet
headers = [
    "Date", "Client Name", "Service Type", "Income (INR)", "Cost of Materials (INR)",
    "Labor Cost (INR)", "Transport Cost (INR)", "Other Expenses (INR)", "Total Expense (INR)", 
    "Profit (INR)"
]

# Apply styles for headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = 18

# Add sample formulas to auto-calculate total expense and profit
for row in range(2, 102):  # Preparing for 100 entries
    ws[f"I{row}"] = f"=SUM(E{row}:H{row})"  # Total Expense = E + F + G + H
    ws[f"J{row}"] = f"=D{row}-I{row}"       # Profit = Income - Total Expense

# Save the workbook
wb.save("CCTV_Service_Profit_Loss_Management.xlsx")
print("Excel file created: CCTV_Service_Profit_Loss_Management.xlsx")
